from __future__ import print_function
from __future__ import unicode_literals
'''
#Google OATH imports
try:
    from googleapiclient.discovery import build
    from httplib2 import Http
    from oauth2client import file, client, tools
except Exception as error:
    print('Please run the following command to install Google API modules:',
          '\n')
    print('pip3 install --upgrade google-api-python-client oauth2client')
    raise error
#End Google module imports
'''
#TODO: Setup workflow for OAuth2 refresh tokens

import os, sys
from os import environ
import datetime
import sys
import requests
import json
import hashlib
from werkzeug import secure_filename
from werkzeug.security import safe_str_cmp
from openpyxl import Workbook, load_workbook
from functools import wraps
from urllib.parse import urlparse, urljoin
from flask import Flask, flash, render_template, request, abort, redirect, url_for, make_response
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_mongoengine import MongoEngine
#Should not import canvas_API_request function. Instead create an endpoint for specific action.
from canvas_module import update_canvas_email, create_canvas_login, canvas_API_request
from canvas_module import enroll_canvas_student, extract_rubric_data, search_students
from users.user_model import User
from hubspot_webhooks.hubspot_webhook_model import Hubspot_Webhook
from assessments.assessment_model import Criterion
from learning_outcomes.learning_outcome_model import Learning_Outcome
from subjects.subject_model import Subject

#Set the default folder for templates
application = Flask(__name__, template_folder='templates')

#Set application secret key to secure against CSRF
application.secret_key = 'super secret key'
application.config['SESSION_TYPE'] = 'filesystem'
application.config['UPLOAD_FOLDER'] = '/uploads'

#Configure mongodb server connection
application.config['MONGODB_SETTINGS'] = {
    'db': 'canvas_integration',
    'host': 'ds125684.mlab.com:25684',
    'username': 'James',
    'password': environ.get('mongoDB_Password'),
    'authentication_source': 'canvas_integration'
}

#Initialise the mongo engine.
db = MongoEngine(application)

#Configure flask-login
login_manager = LoginManager()
login_manager.session_protection = 'strong'

#Redirect to login view when a user has yet to authenticate.
login_manager.login_view = 'login'
login_manager.init_app(application)

#user_loader callback used to load a user from a session ID.
@login_manager.user_loader
def load_user(user_id):
    return User.objects(pk=user_id).first()

def main():
    application.debug = True
    port = int(os.environ.get('PORT', 5000))
    application.run(host='0.0.0.0', port=port)

@application.route('/')
def home():
    return render_template('home.html')

@application.route('/display_cookies', methods=['GET'])
@login_required
def display_cookies():
    for each in request.cookies:
        print(each)
    return redirect(url_for('home'))

@application.route('/login', methods=['GET','POST'])
def login():
    if(request.method == 'POST'):
        try:
            print(request.args)
            username = str(request.values.get('username'))
            password = str(request.values.get('password'))
        except Exception as error:
            raise error
        else:
            #Inform users of username/password constrains.
            if(safe_str_cmp(user.password.encode('utf-8'), password.encode('utf-8'))):
                user = User.authenticate(username, password)
            if user is not None and user.is_authenticated:
                login_status = login_user(user)
                flash('Logged in successfully.')
                next = get_redirect_target()
                return redirect_back('home', next=next)
                # is_safe_url should check if the url is safe for redirects.
                # See http://flask.pocoo.org/snippets/62/ for an example.
            else:
                return redirect(url_for('signup'), code=302)
    else:
        return render_template('login.html')

def is_safe_url(target):
    ref_url = urlparse(request.host_url)
    target_url = urlparse(urljoin(request.host_url, target))
    return target_url.scheme in ('http', 'https') and \
           ref_url.netloc == target_url.netloc

def get_redirect_target():
    for target in request.values.get('next'), request.referrer:
        if not target:
            continue
        if is_safe_url(target):
            return target

def redirect_back(endpoint, **values):
    target = request.args.get('next')
    if not target or not is_safe_url(target):
        target = url_for(endpoint, **values)
    return redirect(target)

@application.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@application.route('/signup', methods=['GET', 'POST'])
def signup():
    if(request.method == 'POST'):
        try:
            username = str(request.form['username'])
            password = str(request.form['password'])
        except Exception as error:
            raise error
        else:
            if username is not "" or password is not "":
                new_user = User.create(username, password)
                return redirect(url_for('login'))
            else:
                return redirect(url_for('signup'))
    else:
        return render_template('signup.html')

def require_hubspot_signature_validation(func):
    #https://developers.hubspot.com/docs/faq/validating-requests-from-hubspot
    #https://developers.hubspot.com/docs/methods/webhooks/webhooks-overview
    @wraps(func)
    def validate_hubspot_response_signature(*args, **kwargs):
        try:
            hubspot_client_secret = environ.get('hubspot_client_secret')
            hubspot_request_signature = request.headers.get('X-HubSpot-Signature')
            request_method = request.method
            request_uri = request.base_url
            request_body = request.get_data(as_text=True)
        except Exception as error:
            raise error
        else:
            hash_string = hubspot_client_secret + request_method + request_uri + request_body
            try:
                encoded_hash_string = hash_string.encode('utf-8')
                request_signature = hashlib.sha256(encoded_hash_string)
            except Exception as error:
                raise error
            else:
                if(hubspot_request_signature == request_signature.hexdigest()):
                    return func(*args, **kwargs)
                else:
                    print('Unauthenticated')
                    #Replace next line when hubspot works
                    return func(*args, **kwargs)
    return validate_hubspot_response_signature

def require_hubspot_access_token(func):
    @wraps(func)
    def update_hubspot_access_token(*args, **kwargs):
        if('hubspot_access_token' not in request.cookies):
            '''
            https://tools.ietf.org/html/rfc6749#section-1.5
            '''
            return redirect(url_for('refresh_access_token'))
        else:
            try:
                token_expiry = current_user.hubspot_access_token_expiry
                last_token_refresh = current_user.last_hubspot_access_token_request
                '''TODO: Implement logic if access token revolked but expiry is
                still valid
                '''
            except Exception as error:
                print('Could not update cookie with user access token, refreshing access token')
            else:
                if(last_hubspot_access_token_request + hubspot_access_expiry > datetime.now()):
                    return redirect(url_for('refresh_access_token'))
                else:
                    return func(*args, **kwargs)
    return update_hubspot_access_token

@application.route('/request_refresh_token', methods=['GET'])
@login_required
def request_refresh_token():
    try:
        code = request.args.get('code')
        client_id = environ.get('hubspot_client_id')
        client_secret = environ.get('hubspot_client_secret')
        '''
        redirect_uri must match the redirect_uri used to intitiate the OAuth
        connection

        '''
        redirect_uri = url_for('request_refresh_token', _external=True,
                               _scheme='https')
    except Exception as error:
        raise error
    else:
        _headers = {
                    'Content-Type': 'application/x-www-form-urlencoded; charset=utf-8'
                   }
        data = {
                'grant_type':'authorization_code', 
                'client_id': client_id,
                'client_secret': client_secret, 
                'redirect_uri': redirect_uri,
                'code': code
               }
        post_request = requests.post(
                                     'https://api.hubapi.com/oauth/v1/token',
                                     headers=_headers, 
                                     data=data
                                    )
        try:
            refresh_token = post_request.json()['refresh_token']
            access_token_expiry = post_request.json()['expires_in']
        except Exception as error:
            raise error
        else:
            User.set_refresh_token(
                                   current_user.id, 
                                   refresh_token,
                                   access_token_expiry
                                  )
            return redirect(url_for('refresh_access_token'))

@application.route('/refresh_access_token', methods=['GET'])
@login_required
def refresh_access_token():
    try:
        client_id = str(environ.get('hubspot_client_id'))
        client_secret = str(environ.get('hubspot_client_secret'))
    except Exception as error:
        print('client_id or client_secret environment variables cannot be found')
        raise error
    else:
        try:
            refresh_token = current_user.refresh_token
        except Exception as error:
            return redirect(url_for('authenticate_hubspot'))
        else:
            endpoint = "https://api.hubapi.com/oauth/v1/token"
            headers = {
                       "Content-Type": "application/x-www-form-urlencoded",
                       "charset": "utf-8"
                      }
            data = {
                    "grant_type": "refresh_token",
                    "client_id": client_id,
                    "client_secret": client_secret,
                    "refresh_token": refresh_token
                   }

            post_request = requests.post(
                                         endpoint,
                                         headers=headers,
                                         data=data
                                        )
            try:
                access_token = post_request.get_json()['access_token']
            except ValueError as error:
                print("Post request response did not contain an access token")
            except Exception as error:
                raise error
            else:
                next = get_redirect_target()
                response = make_response(redirect_back('home'), next=next)
                response.set_cookie('hubspot_access_token', access_token)
                return response

@application.route('/hubspot')
@login_required
def authenticate_hubspot():
    '''
    Hubspot OAuth workflow:

    Direct users to https://app.hubspot.com/oauth/authorize with the following
    query parameters:
        -client_id
        -scope
        -redirect_uri

    They will be prompted to authenticate and authorise the application.

    Users will be redirected to the redirect_uri with a code query parameter.
    
    Use the code above to request access token and refresh token.
    Headers = Content-Type: application/x-www-form-urlencoded;charset=utf-8
    Data:
        -grant_type=authorisation_code
        -client_id
        -client_secret
        -redirect_uri
        -code
    POST https://app.hubspot.com/oauth/v1/token
    '''
    try:
        client_id = environ.get('hubspot_client_id')
        scope = environ.get('hubspot_scopes')
        redirect_uri = url_for('request_refresh_token', _external=True,
                               _scheme='https')
    except Exception as error:
        raise error
    return redirect('https://app.hubspot.com/oauth/authorize?client_id={0}&scope={1}&redirect_uri={2}'.format(client_id,
                                                                                                  scope,
                                                                                                  redirect_uri))

@application.route('/hubspot/workflows', methods=['GET'])
@require_hubspot_access_token
@login_required
def workflows():
    try:
        access_token = request.cookies.get('hubspot_access_token')
    except Exception as error:
        raise error
    else:
        endpoint = 'https://api.hubapi.com/automation/v3/workflows'
        request_headers = {
                           "Content-Type": "application/json",
                           "Authorization": "Bearer " + str(access_token)
                          }
        workflow_request = requests.get(endpoint, headers=request_headers)
        return workflow_request.get_json()

@application.route('/hubspot/workflow_history/<workflow_id>')
@require_hubspot_access_token
@login_required
def workflow_history(workflow_id):
    access_token = request.cookies.get('hubspot_access_token')
    domain = 'https://api.hubapi.com'
    endpoint = '/automation/v3/logevents/workflows/{0}/filter'
    request_url = domain + endpoint.format(workflow_id)
    request_headers = {
                       "Content-Type": "application/json",
                       "Authorization": "Bearer " + str(access_token)
                      }
    request_body = {
                    "types": ["ENROLLED"]
                   }
    try:
        put_request = requests.put(
                             request_url, 
                             headers=request_headers,
                             params=request_body
                            )
        if put_request.status_code == 401:
            return redirect(url_for('authenticate_hubspot'))
        else:
            return put_request.text
    except Exception as error:
        return redirect(url_for('home'))
    
@application.route('/rubric_data', methods=['GET', 'POST'])
@login_required
def rubric_data():
    if(request.method == 'GET'):
        return render_template('rubric_data.html')
    else:
        try:
            course_ID = request.values.get('course_id')
            assessment_ID = request.values.get('assessment_id')
        except Exception as error:
            raise error
        else:
            rubric_data = extract_rubric_data(course_ID, assessment_ID)
            #Save rubric data to the database.
            #map_rubric_data(request.json())
            return rubric_data.text

@application.route('/subjects', methods=['GET', 'POST'])
@login_required
def subjects():
    if(request.method == 'GET'):
        subjects = Subject.read()
        learning_outcomes = json.loads(Learning_Outcome.read())
        return render_template('subjects.html',
                               subjects=subjects,
                               learning_outcomes=learning_outcomes)
    elif(request.method == 'POST'):
        try:
            subject_code = request.form['subject_code_field']
            subject_name = request.form['subject_name_field']
            subject_description = request.form['subject_description_field']
            learning_outcome_ids = request.form.getlist('subject_learning_outcomes_field[]')
        except Exception as error:
            raise error

        subject_learning_outcomes = []
        for each_learning_outcome_id in learning_outcome_ids:
            subject_learning_outcomes.append(Learning_Outcome.index(each_learning_outcome_id))

        subject = Subject(
                          subject_code,
                          subject_name, 
                          subject_description,
                          subject_learning_outcomes
                         ).save()
        return subject.to_json()
    else:
        return abort(405)

@application.route('/learning_outcomes', methods=['GET', 'POST'])
@login_required
def learning_outcomes():
    if(request.method == 'GET'):
        learning_outcomes = json.loads(Learning_Outcome.read())
        return render_template('learning_outcomes.html', 
                               learning_outcomes=learning_outcomes)
    elif(request.method == 'POST'):
        try:
            learning_outcome_name = request.form['learning_outcome_name_field']
            learning_outcome_description = request.form['learning_outcome_description_field']
        except Exception as error:
            raise error
        try:
            learning_outcome = Learning_Outcome(
                             learning_outcome_name,
                             learning_outcome_description
                            ).save()
            return learning_outcome.to_json()
        except Exception as error:
            return abort(500)
    else:
        return abort(405)

@application.route('/criterion', methods=['GET', 'POST'])
@login_required
def criterion():
    if(request.method == 'GET'):
        criterion = json.loads(Criterion.read())
        learning_outcomes = json.loads(Learning_Outcome.read())
        return render_template('criterion.html',
                              criterion=criterion,
                              learning_outcomes=learning_outcomes)
    elif(request.method == 'POST'):
        try:
            criterion_name = request.form['criterion_name_field']
            criterion_description = request.form['criterion_description_field']
            criterion_points = request.form['criterion_points_field']
            criterion_learning_outcomes = request.form.getlist('criterion_learning_outcomes_field[]')
        except Exception as error:
            raise error

        try:
            criterion = Criterion(criterion_name=criterion_name,
                                  criterion_description=criterion_description,
                                  criterion_points=criterion_points,
                                  criterion_learning_outcomes=criterion_learning_outcomes).save()
            return "Success"
        except Exception as error:
            raise error

@application.route('/assessments', methods=['GET', 'POST'])
@login_required
def assessments():
    if(request.method == 'GET'):
        assessments = json.loads(Assessment.read())
        return render_template('assessments.html', 
                               assessments = assessments)

def map_rubric_data(submission_data):
    for each_submission_item in submission_data:
        try:
            submission_ID = each_submission_item['id']
            submission_assignment_ID = each_submission_item['assignment_id']
        except Exception as error:
            raise error
        try:
            submission_rubric_assessment = each_submission_item['rubric_assessment'] 
        except Exception as error:
            pass
        if(submission_rubric_assessment):
            submission = submission_object(submission_ID, submission_assignment_ID,
                                          submission_rubric_assessment)
            submission_grades = []
            for each_criteria in submission.criteria:
                try:
                    learning_outcome = Learning_Outcome(int(each_criteria.id),
                                                        float(each_criteria.points)).save()
                except Exception as error:
                    #Some points are marked blank and cannot be converted. 
                    pass
                submission_grades.append(learning_outcome)
            assessment = Rubric_Assessment.create(each_submission_item['user_id'], 
                                                  Assessment.objects(assessment_id=667),
                                                  submission_grades)
            learning_outcome_count = 0
            grade_total = 0
            for each_learning_outcome in assessment.grades:
                learning_outcome_count = learning_outcome_count + 1
                grade_total = grade_total + 1
                if(grade_total == 21):
                    print(grade_total)
                if(grade_total == 36):
                    print(grade_total)

class submission_object():
    def __init__(self, submission_ID, submission_assessment_ID,
                 submission_rubric_assessment):
        self.criteria = []
        self.id = submission_ID
        self.assessment_ID = submission_assessment_ID
        for key, value in submission_rubric_assessment.items():
            points = None
            comments = None
            try:
                points = value['points']
                comments = value['comments']
            except KeyError as error:
                pass
            except Exception as error:
                raise error
            if(points != None):
                criterion_object = criterion(key, points, comments)
                self.criteria.append(criterion_object)

@application.route('/students', methods=['GET', 'POST'])
@login_required
def student_search():
    if(request.method == 'POST'):
        try:
            search_term = request.form['search_term']
        except Exception as error:
            raise error
        else:
            search_results = search_students(search_term)
            return search_results.text
    else:
        return render_template('student_search.html')

@application.route('/create-account', methods=['POST'])
#@require_hubspot_signature_validation
def create_canvas_account():
    '''
    Docstring
    ---------
    create_account() should only be run in a production environment
    Arguments
    ---------
    student_data(JSON Object):
        Takes a JSON Object that contains firstname, lastname and email
    Returns
    -------
    Account_Creation_Successful(template):
        Returns a template to be rendered by Flask on successful request.
    Note: A course ID will be sent from the webhook as a query paramter. Is this safe?
    '''
    try:
        course_ID = str(request.args.get('course_id'))
        section_ID = str(request.args.get('section_id'))
    except Exception as error:
        print("Could not convert course/section id's to strings")
    else:
        #Validate POST payload
        if not request.is_json:
            return abort(415)
        else:
            #http://flask.pocoo.org/docs/1.0/api/#response-objects
            #Returns None if JSON could not be parsed.
            json_data = request.get_json()
            try:
                Hubspot_Webhook.create(json_data)
            except Exception as error:
                pass
            #Check if JSON data was parsed correctly.
            #Validate JSON Object is dict not array.
            if json_data and isinstance(json_data, dict):
                try:
                    first_name = json_data['properties']['firstname']['value']
                    last_name = json_data['properties']['lastname']['value'] 
                    student_email = json_data['properties']['email']['value']
                    student_name = first_name + " " + last_name
                except KeyError as error:
                    print("Specified JSON fields are not present")
                    return abort(422)
                except Exception as error:
                    print(error)
                else:
                    creation_response = create_canvas_login(student_name, student_email)
                    if(creation_response.status_code == 400):
                        print("The user already exists")
                        students_found = search_students(student_email).json()
                        best_fit_student = students_found[0] or {}
                        print(type(best_fit_student))
                        if isinstance(best_fit_student, dict):
                            try:
                                user_ID = best_fit_student['id']
                            except KeyError as error:
                                print("Specified JSON fields are not present")
                                return abort(422)
                        else:
                            print('students_found is not a json object')
                            return abort(422)
                    elif(creation_response.status_code == 200):
                        user_details = creation_response.json()
                        try:
                            user_ID = user_details['id']
                        except TypeError as error:
                            print("Specified JSON fields are not present")
                            return abort(422)
                        except Exception as error:
                            raise error
                    else:
                        return creation_response.text
                    #Endpoint will return 422 if student_id doesn't exist
                    #return redirect(url_for('enroll_user_in_course', student_id=user_ID, course_id=course_ID, section_id=section_ID))
                    url = 'https://canvas-integration.herokuapp.com/enroll_student'
                    _data = {
                            "student_id": user_ID,
                            "course_id": course_ID,
                            "section_id": section_ID
                            }
                    student_enrollment_request = requests.post(
                                                              url,
                                                              data=_data
                                                             )
                    '''
                    return redirect(url_for('enroll_user_in_course', 
                        student_id=user_ID, 
                        course_id=course_ID, 
                        section_id=section_ID))
                    '''
                    return str(student_enrollment_request.text)
            else:
                flash("Could not parse JSON, Bad Request")
                return abort(400)

@application.route('/enroll_student', methods=['POST'])
def enroll_user_in_course():
    try:
        course_ID = str(request.args.get('course_id'))
        section_ID = str(request.args.get('section_id'))
        student_ID = str(request.args.get('student_id'))
    except Exception as error:
        raise error
    else:
        student_enrollment_request = enroll_canvas_student(student_ID, course_ID, section_ID)
        return student_enrollment_request.text

#TODO: File upload uri with student
#url https://coderacademy/
#uri /users/:id
@application.route('/sis_id_update', methods=['GET', 'POST'])
@login_required
def update_sis_id():
    if(request.method == 'GET'):
        return render_template('sis_id_uploader.html')
    if(request.method == 'POST'):
        # https://openpyxl.readthedocs.io/en/stable/
        if 'File' not in request.files:
            flask("No file uploaded")
            return redirect(url_for(update_sis_id))
        uploaded_file = request.files['File']
        if(uploaded_file.filename == ""):
            flask("No selected file")
            return redirect(url_for(update_sis_id))
        if uploaded_file:
            filename = secure_filename(uploaded_file.filename)
            uploaded_file.save(os.path.join(application.config['UPLOAD_FOLDER'], filename))

        workbook = Workbook()
        excel_document = load_workbook(os.path.join(app.config['UPLOAD_FOLDER'], filename))

        sheet_names_available = excel_document.get_sheet_names()
        print("Available sheets in given file: ", sheet_names_available)

        selected_sheet = excel_document.get_sheet_by_name(sheet_names_available[0])

        cell_range = selected_sheet['C2': 'C96']
        '''
        for each_row in cell_range:
            for cell in each_row:
                student_email = cell.value
                student_number = (student_email).split('@')[0]
                
                user_id = json.loads(search_students(student_email).text)['id'] 
                sis_id = student_number
                
                #GET USERS LOGIN ID.
                domain = 'https://coderacademy.instructure.com'
                endpoint = '/api/v1/users/{0}/logins'.format(user_id)
                user_login = canvas_API_request(
                                                domain + endpoint,
                                                request_method = 'GET'
                                               )
                user_login_id = json.loads(user_login.text)[0]['id']
                endpoint = '/api/v1/accounts/0/logins/{0}'.format(user_login_id)
                user_login_details = canvas_API_request(
                                                        domain + endpoint,
                                                        request_parameters={'login[sis_user_id]':sis_id},
                                                        request_method = 'PUT'
                                                       )
                return user_login_details.text
        '''
#Opens the YAML file at the specified directory and returns the scriptable YAML object.
def get_config(_dir):
    '''
    Arguments
    ---------
    _dir(String):
        Takes a directory method argument used to load configuration.
    Returns
    -------
    file_content:
        Reads the file specified by the _dir string and returns the contents
        using yaml.load(). Alternatively you could use yaml.safe_load().
    '''
    
    file_content = None

    #Checks whether the _dir method argument is a string. isinstance() supports DataTypes that inherit the String base class.
    assert isinstance(_dir, str)
    
    #Check if the directory method argument exists in the current filesystem.
    if os.path.exists(_dir):
        with open(_dir, 'r') as config_file:
            try:
                print('config accessed and read')
                #Load the configuration into a scriptable object.
                file_content = yaml.load(config_file)
            except IOError as error:
                raise error
                sys.exit(0)
            except EOFError as error:
                raise error
                sys.exit(0)
            except ImportError as error:
                print("YAML module could not be imported, please ensure that YAML module has been installed and is in the requirements.txt file")
            except Exception as error:
                raise error
            if(file_content != None):
                return file_content
            else:
                print("Could not load configuration from file but no errors were thrown")
    else:
        print('_dir method argument is not a valid directory in the current filesystem')

def google_request(spreadsheet_ID, range_name, scope):
    '''
    Arguments
    ---------
    spreadsheet_ID(String):
        Takes a string agrument that represents the google spreadsheet
        identifier.
    range_name(String):
        Take a string argument that represents the google sheet ranges to
        retreive data from. Format for the string is as follows:
            '<sheet_name>!<start_range>:<end_range>'
    scope(Google)
    '''
    store = file.Storage('token.json')
    creds = store.get()
    if not creds or creds.invalid:
        flow = client.flow_from_clientsecrets('credentials.json', scope)
        creds = tools.run_flow(flow, store)
    service = build('sheets', 'v4', http=creds.authorize(Http()))

    result = service.spreadsheets().values().get(spreadsheetId=spreadsheet_ID,
                                                range=range_name).execute()
    sheet_data = result.get('values', [])
    
    if not sheet_data:
        sys.exit()
    else:
        return sheet_data

if __name__ == "__main__":
    main()
